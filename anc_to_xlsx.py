#!/usr/bin/env python3

# usage: ./anc_to_xlsx <source.anc> <target.xlsx> [<category_1> ... [<category_n>] ]


import sys
import datetime
import collections
import openpyxl as pxl

# data and language format-specific constants
RECORD_SEPARATOR = "\n"
RECORD_FIELD_SEPARATOR = ";"
DATE_FORMAT = "%Y.%m.%d.%H.%M.%S"
CURRENCY_SYMBOL = "€"
NUMBER_DECIMAL_SEPARATOR = ","
NUMBER_IRRELEVANT_SEPARATOR = "."
#
MISC_SHEET_NAME = "Sonstiges"


def export_string(cell, field_str):
    cell.value = field_str


def export_date(cell, field_str):
    value = datetime.datetime.strptime(field_str, DATE_FORMAT)
    cell.value = value
    cell.number_format = "DD.MM.YYY"


def export_currency(cell, field_str):
    value = float(
        field_str.strip(CURRENCY_SYMBOL)
        .replace(NUMBER_IRRELEVANT_SEPARATOR, "")
        .replace(NUMBER_DECIMAL_SEPARATOR, ".")
        .strip()
    )
    cell.value = value
    cell.number_format = "#,##0.00" + CURRENCY_SYMBOL


field_exporters = collections.defaultdict(lambda: export_string)
field_exporters[1] = export_date
field_exporters[2] = export_currency


def get_category(record_str, categories):
    for category in categories:
        if category in record_str:
            return category
    return None


# source : .anc
# target : .xlsx
def export(source, target, categories):

    # create new workbook
    target_woorkbook = pxl.Workbook()

    # create category sheets
    misc_sheet = target_woorkbook.active
    misc_sheet.title = MISC_SHEET_NAME
    category_sheets = {None: misc_sheet}
    for category in categories:
        category_sheets[category] = target_woorkbook.create_sheet(category)

    # export CSV-style data from source to data_sheet using the field_exporter
    # The c-th field in the r-th record (line) of source becomes the cell of row r and column c.
    with open(source, "r") as source_file:
        for record_str in source_file.read().split(RECORD_SEPARATOR):
            current_category = get_category(record_str, categories)
            current_category_sheet = category_sheets[current_category]
            current_row = current_category_sheet.max_row + 1
            current_column = 1
            for field_str in record_str.split(RECORD_FIELD_SEPARATOR):
                current_cell = current_category_sheet.cell(
                    row=current_row, column=current_column
                )
                try:
                    field_exporters[current_column](current_cell, field_str)
                except Exception as e:
                    print(
                        "WARNING: Export error for cell ["
                        + str(current_column)
                        + ","
                        + str(current_row)
                        + "]:\n"
                        + str(e)
                    )
                current_column += 1
            current_row += 1

    # resize columns
    for sheet in category_sheets.values():
        sheet.column_dimensions["A"].width = 10
        sheet.column_dimensions["B"].width = 15
        sheet.column_dimensions["C"].width = 45

        target_woorkbook.save(filename=target)


if __name__ == "__main__":
    export(source=sys.argv[1], target=sys.argv[2], categories=sys.argv[3:])
